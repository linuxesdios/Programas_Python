from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QAction, QFileDialog
from PyQt5.QtCore import pyqtSignal, Qt
from openpyxl import load_workbook 

class ExcelTableApp(QMainWindow):
	fila_seleccionada = pyqtSignal(list)

	def __init__(self):
		super().__init__()

		self.file_path, _ = QFileDialog.getOpenFileName(None, "Seleccionar archivo Excel", "", "Archivos Excel (*.xlsx *.xls)")
		if not self.file_path:
			QApplication.quit()

		self.initUI()

	def initUI(self):
		self.setGeometry(100, 100, 800, 600)
		self.setWindowTitle('Tabla desde Excel')

		self.central_widget = QWidget()
		self.setCentralWidget(self.central_widget)

		self.layout = QVBoxLayout()

		self.table_widget = QTableWidget()
		self.layout.addWidget(self.table_widget)

		self.central_widget.setLayout(self.layout)

		self.cargar_excel(self.file_path)

		# Conectar señales
		self.table_widget.cellDoubleClicked.connect(self.seleccionar_fila)
		self.table_widget.cellActivated.connect(self.seleccionar_fila)

	def cargar_excel(self, file_path):
		print('Iniciando carga del archivo...')
		try:
			workbook = load_workbook(filename=file_path)
			sheet = workbook.active

			# Establecer el número de filas y columnas en la tabla
			self.table_widget.setRowCount(sheet.max_row)
			self.table_widget.setColumnCount(sheet.max_column)

			# Llenar la tabla con los datos del archivo Excel
			for row_index, row in enumerate(sheet.iter_rows(values_only=True), 0):
				for col_index, cell_value in enumerate(row, 0):
					self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(cell_value)))

			# Establecer las etiquetas de la cabecera horizontal de la tabla
			self.table_widget.setHorizontalHeaderLabels([str(cell.value) for cell in sheet[1]])

			print('Archivo cargado correctamente. Seleccione una fila para continuar.')

		except Exception as e:
			print(f'Error al abrir el archivo Excel: {e}')

	def mostrar_valores_fila(self, item):
		fila = item.row()
		columnas = self.table_widget.columnCount()
		valores_fila = [self.table_widget.item(fila, col).text() for col in range(columnas)]
		print("Valores de la fila seleccionada:", valores_fila)
		self.fila_seleccionada.emit(valores_fila)

	def abrir_ventana_excel(self):
		try:
			self.fila_seleccionada.connect(self.mostrar_valores_fila)
			self.show()
		except Exception as e:
			print(f'Error al abrir la ventana de Excel: {e}')

	def seleccionar_fila(self, row, col):
		# Seleccionar la fila completa
		self.table_widget.selectRow(row)
		# Emitir la señal para mostrar los valores de la fila
		self.mostrar_valores_fila(self.table_widget.item(row, 0))

# Esta función puede ser desde cualquier otro lugar de tu código
def abrir_ventana_excel_desde_otra_funcion():
	return ExcelTableApp()

if __name__ == "__main__":
	app = QApplication(sys.argv)
	GUI = Contratos_App()

	# Llama a la función para abrir la ventana de Excel
	GUI.abrir_ventana()

	GUI.show()
	sys.exit(app.exec_())
